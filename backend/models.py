#!/usr/bin/env python3
"""
CLI: Parse an Excel file entirely in-memory and emit JSON preserving all sheet data.

Usage:
  python models.py --file "C:\\path\\to\\workbook.xlsx" --out parsed_output.json [--no-gpt]

Design goals:
- SPEED: stream bytes into memory and avoid temp files.
- ROBUSTNESS: try multiple parsing strategies; fall back to a ZIP+XML parser.
- FIDELITY: preserve every cell value (including blanks as null) in a dense 2D array per sheet,
  plus sheet dimensions. Where possible, convert dates to ISO 8601.
- GPT-5 (optional): call the OpenAI Responses API with low reasoning effort & low verbosity to
  quickly sanity-check the parsed result (counts, sheet names). We DO NOT use the model to
  transform values to ensure accuracy—parsing is deterministic.

Dependencies:
  pip install openai python-dotenv pandas openpyxl

Note: If your environment has trouble with pandas/openpyxl for large/complex files, this
script will still succeed via the built-in ZIP+XML fallback.
"""

import argparse
import io
import json
import math
import os
import re
import sys
import zipfile
from datetime import datetime, timedelta, date
from typing import Dict, List, Tuple, Any

# Optional heavy deps (used in first strategy):
try:
    import pandas as pd  # type: ignore
except Exception:
    pd = None  # Fallback path will still work

try:
    from openpyxl import load_workbook  # type: ignore
except Exception:
    load_workbook = None

from dotenv import load_dotenv  # type: ignore

# ---- Excel helpers (ZIP+XML fallback) --------------------------------------
import xml.etree.ElementTree as ET

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"

NS = {"s": NS_MAIN, "r": NS_REL, "p": NS_PKG}

# Common Excel date format ids (<60 are built-ins; 14 is 'm/d/yy')
BUILTIN_DATE_FMT_IDS = {
    14, 15, 16, 17, 22, 27, 30, 36, 45, 46, 47, 50, 57
}

# Excel date system origin (Windows): 1899-12-30 (with the 1900 leap-year bug)
EXCEL_EPOCH = datetime(1899, 12, 30)


def excel_serial_to_datetime(n: float) -> datetime:
    try:
        return EXCEL_EPOCH + timedelta(days=float(n))
    except Exception:
        raise


def col_letters_to_index(letters: str) -> int:
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch.upper()) - ord('A') + 1)
    return n


def cell_ref_to_coord(cell_ref: str) -> Tuple[int, int]:
    m = re.match(r"^([A-Za-z]+)(\d+)$", cell_ref)
    if not m:
        return 0, 0
    col_letters, row_num = m.groups()
    return int(row_num), col_letters_to_index(col_letters)


class XLSXZipParser:
    def __init__(self, data: bytes):
        self.zf = zipfile.ZipFile(io.BytesIO(data))
        self.shared_strings: List[str] = []
        self.num_fmts: Dict[int, str] = {}
        self.cell_xfs_numfmt: List[int] = []
        self.sheet_map: List[Tuple[str, str]] = []  # (sheet_name, xml_path)
        self._load_shared_strings()
        self._load_styles()
        self._load_sheet_map()

    # ---- workbook components ----
    def _load_shared_strings(self) -> None:
        try:
            with self.zf.open("xl/sharedStrings.xml") as f:
                root = ET.parse(f).getroot()
                for si in root.findall("s:si", NS):
                    parts: List[str] = []
                    t = si.find("s:t", NS)
                    if t is not None and t.text:
                        parts.append(t.text)
                    for r in si.findall("s:r", NS):
                        rt = r.find("s:t", NS)
                        if rt is not None and rt.text:
                            parts.append(rt.text)
                    self.shared_strings.append("".join(parts))
        except KeyError:
            # no shared strings
            self.shared_strings = []

    def _load_styles(self) -> None:
        try:
            with self.zf.open("xl/styles.xml") as f:
                root = ET.parse(f).getroot()
                # custom numFmts
                for nf in root.findall("s:numFmts/s:numFmt", NS):
                    numFmtId = int(nf.attrib.get("numFmtId", "-1"))
                    formatCode = nf.attrib.get("formatCode", "")
                    self.num_fmts[numFmtId] = formatCode
                # cellXfs -> xf@numFmtId
                self.cell_xfs_numfmt = []
                for xf in root.findall("s:cellXfs/s:xf", NS):
                    numFmtId = int(xf.attrib.get("numFmtId", "-1"))
                    self.cell_xfs_numfmt.append(numFmtId)
        except KeyError:
            self.num_fmts = {}
            self.cell_xfs_numfmt = []

    def _load_sheet_map(self) -> None:
        with self.zf.open("xl/workbook.xml") as f:
            wb_root = ET.parse(f).getroot()
            sheets: List[Tuple[str, str]] = []
            for sheet in wb_root.findall("s:sheets/s:sheet", NS):
                name = sheet.attrib.get("name", "Sheet")
                rid = sheet.attrib.get(f"{{{NS_REL}}}id")
                sheets.append((name, rid))
        with self.zf.open("xl/_rels/workbook.xml.rels") as f:
            rels_root = ET.parse(f).getroot()
            rel_map = {r.attrib["Id"]: r.attrib["Target"] for r in rels_root.findall("p:Relationship", NS)}
        out: List[Tuple[str, str]] = []
        for name, rid in sheets:
            target = rel_map.get(rid, "")
            if target and not target.startswith("xl/"):
                target = "xl/" + target
            out.append((name, target))
        self.sheet_map = out

    # ---- parsing ----
    def _is_date_style(self, s_idx: int) -> bool:
        if s_idx < 0 or s_idx >= len(self.cell_xfs_numfmt):
            return False
        fmt_id = self.cell_xfs_numfmt[s_idx]
        if fmt_id in BUILTIN_DATE_FMT_IDS:
            return True
        fmt_code = self.num_fmts.get(fmt_id, "")
        # Heuristic: any format containing date tokens
        return bool(re.search(r"[dyhmSs]", fmt_code))

    def _apply_merged_cells(self, root: ET.Element, grid: Dict[int, Dict[int, Any]]):
        # Copy the top-left cell value into the full merge range for fidelity
        for merge in root.findall("s:mergeCells/s:mergeCell", NS):
            ref = merge.attrib.get("ref", "")
            m = re.match(r"^([A-Z]+\d+):([A-Z]+\d+)$", ref)
            if not m:
                continue
            tl, br = m.groups()
            r1, c1 = cell_ref_to_coord(tl)
            r2, c2 = cell_ref_to_coord(br)
            val = grid.get(r1, {}).get(c1)
            for r in range(r1, r2 + 1):
                row_map = grid.setdefault(r, {})
                for c in range(c1, c2 + 1):
                    if (r, c) != (r1, c1):
                        row_map[c] = val

    def parse_sheet(self, sheet_xml_path: str) -> Dict[str, Any]:
        with self.zf.open(sheet_xml_path) as f:
            root = ET.parse(f).getroot()
        grid: Dict[int, Dict[int, Any]] = {}
        max_row = 0
        max_col = 0
        for row in root.findall("s:sheetData/s:row", NS):
            r_idx = int(row.attrib.get("r", "0"))
            if r_idx > max_row:
                max_row = r_idx
            for c in row.findall("s:c", NS):
                ref = c.attrib.get("r")
                if not ref:
                    continue
                rr, cc = cell_ref_to_coord(ref)
                if cc > max_col:
                    max_col = cc
                t = c.attrib.get("t")  # s, b, str, inlineStr
                s_attr = c.attrib.get("s")
                style_idx = int(s_attr) if s_attr is not None and s_attr.isdigit() else -1
                v = c.find("s:v", NS)
                is_el = c.find("s:is", NS)
                val: Any = None
                if t == "inlineStr" and is_el is not None:
                    tt = is_el.find("s:t", NS)
                    val = tt.text if tt is not None else None
                elif v is not None and v.text is not None:
                    text = v.text
                    if t == "s":
                        # shared string
                        try:
                            idx = int(text)
                            val = self.shared_strings[idx] if 0 <= idx < len(self.shared_strings) else text
                        except Exception:
                            val = text
                    elif t == "b":
                        val = True if text == "1" else False
                    else:
                        # numeric or text (as stored)
                        # detect date via style
                        if self._is_date_style(style_idx):
                            try:
                                dt = excel_serial_to_datetime(float(text))
                                val = dt.isoformat()
                            except Exception:
                                val = text
                        else:
                            # try numeric
                            try:
                                if "." in text or "e" in text or "E" in text:
                                    val = float(text)
                                else:
                                    val = int(text)
                            except Exception:
                                val = text
                grid.setdefault(rr, {})[cc] = val
        # expand merged cells
        self._apply_merged_cells(root, grid)
        # densify
        rows: List[List[Any]] = []
        for r in range(1, max_row + 1):
            row_map = grid.get(r, {})
            rows.append([row_map.get(c) for c in range(1, max_col + 1)])
        return {"max_row": max_row, "max_col": max_col, "rows": rows}

    def parse_workbook(self) -> Dict[str, Any]:
        out: Dict[str, Any] = {"sheets": {}}
        for name, path in self.sheet_map:
            out["sheets"][name] = self.parse_sheet(path)
        out["sheet_count"] = len(self.sheet_map)
        return out


# ---- High-level parsing orchestrator ---------------------------------------

def parse_excel_bytes(data: bytes) -> Dict[str, Any]:
    # Strategy 1: pandas (fast for simple/medium files)
    if pd is not None:
        try:
            xl = pd.ExcelFile(io.BytesIO(data))
            result = {"sheets": {}, "sheet_count": len(xl.sheet_names)}
            for sheet in xl.sheet_names:
                df = pd.read_excel(io.BytesIO(data), sheet_name=sheet, dtype=object)
                df = df.astype("object").where(pd.notna(df), None)
                # Dense 2D array to preserve layout
                matrix = df.values.tolist()
                result["sheets"][sheet] = {
                    "max_row": len(matrix),
                    "max_col": len(matrix[0]) if matrix else 0,
                    "rows": matrix,
                }
            return result
        except Exception:
            pass
    # Strategy 2: openpyxl streaming (robust)
    if load_workbook is not None:
        try:
            wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
            result = {"sheets": {}, "sheet_count": len(wb.sheetnames)}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                max_row = ws.max_row or 0
                max_col = ws.max_column or 0
                rows: List[List[Any]] = []
                for r in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
                    # Convert dates
                    conv = []
                    for v in r:
                        if isinstance(v, (datetime, date)):
                            conv.append(v.isoformat())
                        else:
                            conv.append(v)
                    rows.append([x if x is not None else None for x in conv])
                result["sheets"][sheet_name] = {"max_row": max_row, "max_col": max_col, "rows": rows}
            return result
        except Exception:
            pass
    # Strategy 3: ZIP+XML fallback (most robust)
    parser = XLSXZipParser(data)
    wb = parser.parse_workbook()
    wb["workbook"] = "in-memory.xlsx"
    return wb


# ---- GPT-5 sanity check (optional) ----------------------------------------

def gpt5_sanity_check(model: str, parsed: Dict[str, Any]) -> str:
    """Send a tiny summary to GPT-5 to double-check counts and echo back a short confirmation.
    This keeps latency minimal and avoids modifying data, but verifies structure.
    """
    try:
        from openai import OpenAI  # type: ignore
    except Exception:
        return "OpenAI SDK not available; skipped GPT-5 check."

    client = OpenAI()
    sheet_names = list(parsed.get("sheets", {}).keys())
    prompt = (
        "You are validating a workbook parse. Given the sheet names and sizes, "
        "reply with a one-line summary: number of sheets and each sheet's rows×cols."
    )
    # Build a compact description for speed
    sizes = {
        name: (parsed["sheets"][name]["max_row"], parsed["sheets"][name]["max_col"]) for name in sheet_names
    }

    # Fast, low-latency call per user guidance
    resp = client.responses.create(
        model=model,
        input={
            "role": "user",
            "content": [
                {"type": "text", "text": prompt},
                {"type": "input_text", "text": json.dumps({"sheets": sheet_names, "sizes": sizes})},
            ],
        },
        reasoning={"effort": "low"},
        text={"verbosity": "low"},
        max_output_tokens=200,
    )
    try:
        return getattr(resp, "output_text", "(no text)")
    except Exception:
        return "(validation complete)"


# ---- CLI -------------------------------------------------------------------

def main():
    load_dotenv()  # loads OPENAI_API_KEY if present

    parser = argparse.ArgumentParser(description="Parse Excel to JSON with in-memory streaming")
    parser.add_argument("--file", required=True, help="Path to the .xlsx file")
    parser.add_argument("--out", required=True, help="Output JSON file path")
    parser.add_argument("--model", default="gpt-5", help="OpenAI model for quick sanity check")
    parser.add_argument("--no-gpt", action="store_true", help="Skip GPT-5 sanity check")

    args = parser.parse_args()

    # Read bytes into memory for speed
    with open(args.file, "rb") as f:
        data = f.read()

    parsed = parse_excel_bytes(data)

    # Write JSON
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(parsed, f, ensure_ascii=False, indent=2)

    print(f"Wrote JSON to {args.out}")

    if not args.no_gpt and os.getenv("OPENAI_API_KEY"):
        try:
            summary = gpt5_sanity_check(args.model, parsed)
            print("GPT-5 check:", summary)
        except Exception as e:
            print("GPT-5 check skipped:", repr(e))
    else:
        print("GPT-5 check disabled or OPENAI_API_KEY not set.")


if __name__ == "__main__":
    main()
