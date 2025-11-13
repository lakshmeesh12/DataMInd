# ingest.py - ENHANCED & ROBUST VERSION
import io
import json
import uuid
import datetime
from datetime import timedelta
from typing import Dict, Any, List, Optional, Tuple
from fastapi import UploadFile
from db.mongo.config import db  # ← Use shared db
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging
import asyncio
import re
import os
from neo4j import AsyncGraphDatabase
from dotenv import load_dotenv
import openai
from db.qdrant.config import (
    qdrant_client,
    QDRANT_COLLECTION,
    QDRANT_COLLECTION_1,        # ← NEW: excel_documents collection
    EMBEDDING_MODEL,
    init_qdrant_excel_collection  # ← NEW: initializer
)
from qdrant_client.http.models import PointStruct

load_dotenv()

logger = logging.getLogger(__name__)

# === Neo4j Configuration ===
NEO4J_URI = os.getenv("NEO4J_URI", "bolt://localhost:7687")
NEO4J_USER = os.getenv("NEO4J_USER", "neo4j")
NEO4J_PASSWORD = os.getenv("NEO4J_PASSWORD", "test12345")

# Initialize Neo4j driver
neo4j_driver = AsyncGraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USER, NEO4J_PASSWORD))

# === Optional fast parsers ===
try:
    import pandas as pd
except Exception:
    pd = None
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

# === ZIP+XML Fallback Parser ===
import zipfile
import xml.etree.ElementTree as ET

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"
NS = {"s": NS_MAIN, "r": NS_REL, "p": NS_PKG}

BUILTIN_DATE_FMT_IDS = {14, 15, 16, 17, 22, 27, 30, 36, 45, 46, 47, 50, 57}
EXCEL_EPOCH = datetime.datetime(1899, 12, 30)

# Global executor
executor = ThreadPoolExecutor(max_workers=8)

# ==================== NORMALIZATION RULES ====================

# Currency patterns and mappings
CURRENCY_PATTERNS = [
    (r'₹\s*([0-9,]+\.?\d*)', 'INR'),
    (r'Rs\.?\s*([0-9,]+\.?\d*)', 'INR'),
    (r'\$\s*([0-9,]+\.?\d*)', 'USD'),
    (r'€\s*([0-9,]+\.?\d*)', 'EUR'),
    (r'£\s*([0-9,]+\.?\d*)', 'GBP'),
]

# Unit normalization mappings
UNIT_MAPPINGS = {
    # Area
    'sqm': 'm²', 'sq.m': 'm²', 'sq m': 'm²', 'square meter': 'm²', 'square metre': 'm²',
    'sqft': 'ft²', 'sq.ft': 'ft²', 'sq ft': 'ft²', 'square feet': 'ft²',
    
    # Length
    'rmt': 'm', 'r.m.t': 'm', 'running meter': 'm', 'running metre': 'm',
    'rm': 'm', 'r.m': 'm',
    'km': 'km', 'meter': 'm', 'metre': 'm', 'mm': 'mm', 'cm': 'cm',
    'feet': 'ft', 'foot': 'ft', 'inch': 'in', 'inches': 'in',
    
    # Volume
    'cum': 'm³', 'cu.m': 'm³', 'cu m': 'm³', 'cubic meter': 'm³', 'cubic metre': 'm³',
    'cuft': 'ft³', 'cu.ft': 'ft³', 'cu ft': 'ft³', 'cubic feet': 'ft³',
    'ltr': 'L', 'litre': 'L', 'liter': 'L',
    
    # Weight
    'kg': 'kg', 'kilogram': 'kg', 'kgs': 'kg',
    'gm': 'g', 'gram': 'g', 'gms': 'g',
    'ton': 't', 'tonne': 't', 'mt': 't', 'metric ton': 't',
    
    # Quantity
    'nos': 'nos', 'no': 'nos', 'number': 'nos', 'numbers': 'nos',
    'pcs': 'pcs', 'piece': 'pcs', 'pieces': 'pcs',
    'set': 'set', 'sets': 'set',
    'bag': 'bag', 'bags': 'bag',
    
    # Others
    'ls': 'LS', 'lump sum': 'LS', 'lumpsum': 'LS',
    'each': 'each', 'ea': 'each',
}

# Common construction item patterns
ITEM_PATTERNS = [
    r'(?i)\b(rcc|pcc|concrete|cement)\b',
    r'(?i)\b(brick|block|masonry)\b',
    r'(?i)\b(steel|reinforcement|rebar|tmt)\b',
    r'(?i)\b(painting|plastering|finishing)\b',
    r'(?i)\b(door|window|frame)\b',
    r'(?i)\b(tile|flooring|marble|granite)\b',
    r'(?i)\b(electrical|plumbing|hvac)\b',
    r'(?i)\b(excavation|earthwork|filling)\b',
    r'(?i)\b(wall|ceiling|roof|slab)\b',
]

# ==================== NORMALIZATION FUNCTIONS ====================

def normalize_currency(value: Any) -> Optional[Dict[str, Any]]:
    """Extract and normalize currency values."""
    if not isinstance(value, str):
        return None
    
    for pattern, currency_code in CURRENCY_PATTERNS:
        match = re.search(pattern, value)
        if match:
            amount_str = match.group(1).replace(',', '')
            try:
                amount = float(amount_str)
                return {
                    'amount': amount,
                    'currency': currency_code,
                    'original': value
                }
            except ValueError:
                continue
    return None

def normalize_unit(value: Any) -> Optional[Dict[str, Any]]:
    """Extract and normalize units."""
    if not isinstance(value, str):
        return None
    
    # Look for number followed by unit
    pattern = r'([0-9,]+\.?\d*)\s*([a-zA-Z\.]+)'
    match = re.search(pattern, value)
    
    if match:
        quantity_str = match.group(1).replace(',', '')
        unit_str = match.group(2).lower().strip('.')
        
        try:
            quantity = float(quantity_str)
            normalized_unit = UNIT_MAPPINGS.get(unit_str, unit_str)
            
            return {
                'quantity': quantity,
                'unit': normalized_unit,
                'original': value
            }
        except ValueError:
            pass
    
    # Check if it's just a unit without number
    for variant, standard in UNIT_MAPPINGS.items():
        if re.search(r'\b' + re.escape(variant) + r'\b', value.lower()):
            return {
                'unit': standard,
                'original': value
            }
    
    return None

def normalize_item(value: Any) -> Optional[Dict[str, Any]]:
    """Extract and normalize construction items."""
    if not isinstance(value, str):
        return None
    
    matches = []
    for pattern in ITEM_PATTERNS:
        found = re.findall(pattern, value)
        matches.extend(found)
    
    if matches:
        return {
            'item_type': matches[0].lower(),
            'keywords': list(set(m.lower() for m in matches)),
            'description': value,
            'original': value
        }
    
    return None

def enrich_cell_value(value: Any) -> Dict[str, Any]:
    """Enrich a cell value with normalized metadata."""
    enriched = {
        'raw': value,
        'type': type(value).__name__
    }
    
    if value is None:
        enriched['type'] = 'null'
        return enriched
    
    # Try currency normalization
    currency = normalize_currency(value)
    if currency:
        enriched['currency'] = currency
        enriched['type'] = 'currency'
    
    # Try unit normalization
    unit = normalize_unit(value)
    if unit:
        enriched['unit'] = unit
        if enriched['type'] == 'str':
            enriched['type'] = 'quantity'
    
    # Try item normalization
    item = normalize_item(value)
    if item:
        enriched['item'] = item
        if enriched['type'] == 'str':
            enriched['type'] = 'item_description'
    
    return enriched

# ==================== TABLE & REGION DETECTION ====================

def detect_header_row(rows: List[List[Any]]) -> Optional[int]:
    """Detect the header row index based on content patterns."""
    for idx, row in enumerate(rows[:20]):  # Check first 20 rows
        if not row:
            continue
        
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if len(non_empty) < 2:
            continue
        
        # Check for common header patterns
        header_indicators = ['item', 'description', 'quantity', 'rate', 'amount', 
                           'unit', 'total', 'no', 's.no', 'sr', 'code']
        
        text_content = ' '.join(str(c).lower() for c in non_empty)
        matches = sum(1 for indicator in header_indicators if indicator in text_content)
        
        if matches >= 2:
            return idx
    
    return None

def detect_table_regions(rows: List[List[Any]], max_col: int) -> List[Dict[str, Any]]:
    """Detect distinct table regions in the sheet."""
    regions = []
    current_region = None
    empty_row_count = 0
    
    for row_idx, row in enumerate(rows):
        non_empty_cells = sum(1 for c in row if c is not None and str(c).strip())
        
        # Start new region if we have content after empty rows
        if non_empty_cells >= 2:
            empty_row_count = 0
            
            if current_region is None:
                current_region = {
                    'start_row': row_idx,
                    'end_row': row_idx,
                    'max_col': max_col,
                    'header_row': None,
                    'data_rows': []
                }
            else:
                current_region['end_row'] = row_idx
        else:
            empty_row_count += 1
            
            # End region if we have 3+ consecutive empty rows
            if empty_row_count >= 3 and current_region is not None:
                # Try to detect header in this region
                region_rows = rows[current_region['start_row']:current_region['end_row']+1]
                header_idx = detect_header_row(region_rows)
                if header_idx is not None:
                    current_region['header_row'] = current_region['start_row'] + header_idx
                
                regions.append(current_region)
                current_region = None
    
    # Add last region if exists
    if current_region is not None:
        region_rows = rows[current_region['start_row']:current_region['end_row']+1]
        header_idx = detect_header_row(region_rows)
        if header_idx is not None:
            current_region['header_row'] = current_region['start_row'] + header_idx
        regions.append(current_region)
    
    return regions

def extract_structured_tables(rows: List[List[Any]], regions: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Extract structured table data from detected regions."""
    tables = []
    
    for region_idx, region in enumerate(regions):
        start = region['start_row']
        end = region['end_row']
        header_row_idx = region['header_row']
        
        table = {
            'region_id': region_idx,
            'start_row': start,
            'end_row': end,
            'headers': None,
            'rows': [],
            'enriched_rows': []
        }
        
        # Extract headers
        if header_row_idx is not None:
            headers = rows[header_row_idx]
            table['headers'] = [str(h).strip() if h is not None else f"col_{i}" 
                               for i, h in enumerate(headers)]
            data_start = header_row_idx + 1
        else:
            # Generate default headers
            table['headers'] = [f"col_{i}" for i in range(region['max_col'])]
            data_start = start
        
        # Extract data rows
        for row_idx in range(data_start, end + 1):
            if row_idx >= len(rows):
                break
            
            row = rows[row_idx]
            non_empty = sum(1 for c in row if c is not None and str(c).strip())
            
            if non_empty >= 1:
                # Store raw row
                table['rows'].append(row)
                
                # Store enriched row
                enriched_row = [enrich_cell_value(cell) for cell in row]
                table['enriched_rows'].append(enriched_row)
        
        if table['rows']:
            tables.append(table)
    
    return tables

# ==================== SEMANTIC CHUNKING ====================

def create_semantic_chunks(tables: List[Dict[str, Any]], sheet_name: str, filename: str) -> List[Dict[str, Any]]:
    """Create semantic chunks from extracted tables."""
    chunks = []
    
    for table in tables:
        # Create a chunk for each table
        chunk = {
            'chunk_id': str(uuid.uuid4()),
            'filename': filename,
            'sheet_name': sheet_name,
            'chunk_type': 'table',
            'region_id': table['region_id'],
            'row_range': {
                'start': table['start_row'],
                'end': table['end_row']
            },
            'headers': table['headers'],
            'row_count': len(table['rows']),
            'metadata': {
                'has_currency': False,
                'has_quantities': False,
                'has_items': False,
                'currency_columns': [],
                'quantity_columns': [],
                'item_columns': []
            }
        }
        
        # Analyze enriched data for metadata
        for col_idx, header in enumerate(table['headers']):
            has_currency = False
            has_quantity = False
            has_item = False
            
            for enriched_row in table['enriched_rows']:
                if col_idx < len(enriched_row):
                    cell = enriched_row[col_idx]
                    if 'currency' in cell:
                        has_currency = True
                    if 'unit' in cell:
                        has_quantity = True
                    if 'item' in cell:
                        has_item = True
            
            if has_currency:
                chunk['metadata']['currency_columns'].append(header)
                chunk['metadata']['has_currency'] = True
            if has_quantity:
                chunk['metadata']['quantity_columns'].append(header)
                chunk['metadata']['has_quantities'] = True
            if has_item:
                chunk['metadata']['item_columns'].append(header)
                chunk['metadata']['has_items'] = True
        
        # Store sample data (first 5 rows)
        chunk['sample_data'] = {
            'raw': table['rows'][:5],
            'enriched': table['enriched_rows'][:5]
        }
        
        chunks.append(chunk)
    
    return chunks

# ==================== DOCUMENT STRUCTURE ANALYSIS (NEW) ====================

def analyze_cell_visual_properties(row_idx: int, col_idx: int, cell_value: Any,
                                   sheet_rows: List[List[Any]]) -> Dict[str, Any]:
    """
    Analyze visual properties of a cell to determine its role in document structure.
    This works WITHOUT needing Excel formatting - uses heuristics.
    """
    if cell_value is None or str(cell_value).strip() == "":
        return {"role": "empty", "confidence": 1.0}
   
    cell_str = str(cell_value).strip()
    properties = {
        "role": "data",
        "confidence": 0.5,
        "is_bold_likely": False,
        "is_header_likely": False,
        "is_section_likely": False,
        "span_estimate": 1,
        "level": 0
    }
   
    # Check if cell is likely a major header (project title, document title)
    is_upper = cell_str.isupper()
    is_early_row = row_idx < 5
    is_long = len(cell_str) > 20
   
    # Check if this cell has many empty cells after it (suggests it spans multiple columns)
    empty_after = 0
    for c in range(col_idx + 1, min(col_idx + 6, len(sheet_rows[row_idx]))):
        if sheet_rows[row_idx][c] is None or str(sheet_rows[row_idx][c]).strip() == "":
            empty_after += 1
        else:
            break
   
    if is_upper and is_early_row:
        properties["role"] = "document_header"
        properties["confidence"] = 0.9
        properties["is_bold_likely"] = True
        properties["span_estimate"] = empty_after + 1
        properties["level"] = 0
        return properties
   
    # Check if it's a section header (like "1 CIVIL WORKS", "2.01 PARTITIONS")
    section_pattern = r'^(\d+\.?\d*\.?\d*)\s+([A-Z\s&/]+)$'
    match = re.match(section_pattern, cell_str)
    if match:
        code, title = match.groups()
        level = code.count('.') + 1
        properties["role"] = "section_header"
        properties["confidence"] = 0.95
        properties["is_bold_likely"] = True
        properties["section_code"] = code
        properties["section_title"] = title.strip()
        properties["level"] = level
        properties["span_estimate"] = empty_after + 1
        return properties
   
    # Check if it's a subsection or category (often has leading spaces or indentation)
    if cell_str.startswith(' ') and len(cell_str.strip()) > 0:
        properties["role"] = "subsection"
        properties["confidence"] = 0.7
        properties["level"] = 2
        return properties
   
    # Check if it's a table header row (common keywords)
    header_keywords = ['item', 'description', 'quantity', 'rate', 'amount', 'unit',
                      'make', 'brand', 'specification', 'code', 'sr', 's.no']
    if any(kw in cell_str.lower() for kw in header_keywords):
        properties["role"] = "table_header"
        properties["confidence"] = 0.8
        properties["is_header_likely"] = True
        return properties
   
    # Check if it's likely a data cell with special meaning
    if col_idx > 0:
        prev_cell = sheet_rows[row_idx][col_idx - 1]
        if prev_cell and len(str(prev_cell).strip()) > 10:
            properties["role"] = "related_value"
            properties["confidence"] = 0.6
   
    return properties


def build_document_hierarchy(sheet_rows: List[List[Any]], sheet_name: str) -> Dict[str, Any]:
    """
    Build a hierarchical representation of the document structure.
    This captures: Document Header → Sections → Subsections → Data
    """
    hierarchy = {
        "document_metadata": {},
        "sections": [],
        "current_section": None,
        "current_subsection": None,
        "flat_structure": []  # For search: every element with full context path
    }
   
    context_stack = []  # Track current position in hierarchy
   
    for row_idx, row in enumerate(sheet_rows):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
       
        for col_idx, cell in enumerate(row):
            if cell is None or str(cell).strip() == "":
                continue
           
            props = analyze_cell_visual_properties(row_idx, col_idx, cell, sheet_rows)
           
            if props["role"] == "document_header":
                header_text = str(cell).strip()
                hierarchy["document_metadata"][f"header_{len(hierarchy['document_metadata'])}"] = header_text
                context_stack = [header_text]
               
                hierarchy["flat_structure"].append({
                    "type": "document_header",
                    "text": header_text,
                    "row": row_idx,
                    "col": col_idx,
                    "context_path": [header_text],
                    "sheet": sheet_name
                })
           
            elif props["role"] == "section_header":
                section = {
                    "code": props.get("section_code"),
                    "title": props.get("section_title"),
                    "level": props.get("level", 1),
                    "start_row": row_idx,
                    "subsections": [],
                    "items": []
                }
                hierarchy["sections"].append(section)
                hierarchy["current_section"] = section
                hierarchy["current_subsection"] = None
               
                context_stack = list(hierarchy["document_metadata"].values()) + [section["title"]]
               
                hierarchy["flat_structure"].append({
                    "type": "section_header",
                    "text": f"{section['code']} {section['title']}",
                    "code": section["code"],
                    "title": section["title"],
                    "row": row_idx,
                    "col": col_idx,
                    "context_path": context_stack.copy(),
                    "sheet": sheet_name
                })
           
            elif props["role"] == "subsection":
                if hierarchy["current_section"]:
                    subsection = {
                        "title": str(cell).strip(),
                        "row": row_idx,
                        "items": []
                    }
                    hierarchy["current_section"]["subsections"].append(subsection)
                    hierarchy["current_subsection"] = subsection
                   
                    context_stack = (list(hierarchy["document_metadata"].values()) +
                                   [hierarchy["current_section"]["title"], subsection["title"]])
                   
                    hierarchy["flat_structure"].append({
                        "type": "subsection",
                        "text": subsection["title"],
                        "row": row_idx,
                        "col": col_idx,
                        "context_path": context_stack.copy(),
                        "sheet": sheet_name
                    })
           
            elif props["role"] in ["data", "related_value"]:
                item = {
                    "text": str(cell).strip(),
                    "row": row_idx,
                    "col": col_idx,
                    "type": props["role"]
                }
               
                row_context = []
                for c_idx, c_val in enumerate(row):
                    if c_val and str(c_val).strip() and c_idx != col_idx:
                        row_context.append(str(c_val).strip())
               
                item["row_context"] = row_context
               
                if hierarchy["current_subsection"]:
                    hierarchy["current_subsection"]["items"].append(item)
                elif hierarchy["current_section"]:
                    hierarchy["current_section"]["items"].append(item)
               
                hierarchy["flat_structure"].append({
                    "type": "data_item",
                    "text": item["text"],
                    "row": row_idx,
                    "col": col_idx,
                    "context_path": context_stack.copy(),
                    "row_context": row_context,
                    "sheet": sheet_name
                })
   
    return hierarchy


def create_context_aware_chunks(hierarchy: Dict[str, Any], filename: str) -> List[Dict[str, Any]]:
    """
    Create chunks that preserve document structure and context.
    Each chunk knows its position in the document hierarchy.
    """
    chunks = []
   
    doc_headers = list(hierarchy["document_metadata"].values())
    doc_context = " | ".join(doc_headers) if doc_headers else filename
   
    for section in hierarchy["sections"]:
        section_text = f"{section['code']} {section['title']}"
        section_items = []
       
        for item in section.get("items", []):
            section_items.append({
                "text": item["text"],
                "row": item["row"],
                "col": item["col"],
                "row_context": item.get("row_context", [])
            })
       
        for subsection in section.get("subsections", []):
            subsection_title = subsection["title"]
            for item in subsection.get("items", []):
                section_items.append({
                    "text": item["text"],
                    "row": item["row"],
                    "col": item["col"],
                    "row_context": item.get("row_context", []),
                    "subsection": subsection_title
                })
       
        searchable_text = f"Document: {doc_context}\n"
        searchable_text += f"Section: {section_text}\n"
        searchable_text += f"Items:\n"
       
        for item in section_items[:50]:
            subsec = f"[{item.get('subsection', '')}] " if item.get("subsection") else ""
            context = f" ({', '.join(item['row_context'][:3])})" if item["row_context"] else ""
            searchable_text += f"- {subsec}{item['text']}{context}\n"
       
        chunk = {
            "chunk_id": str(uuid.uuid4()),
            "filename": filename,
            "sheet_name": hierarchy["flat_structure"][0]["sheet"] if hierarchy["flat_structure"] else "Unknown",
            "chunk_type": "hierarchical_section",
            "document_context": doc_context,
            "section_code": section["code"],
            "section_title": section["title"],
            "section_level": section["level"],
            "item_count": len(section_items),
            "items": section_items,
            "subsections": [s["title"] for s in section.get("subsections", [])],
            "searchable_text": searchable_text,
            "context_path": [doc_context, section_text]
        }
        chunks.append(chunk)
   
    # Document overview chunk
    overview_chunk = {
        "chunk_id": str(uuid.uuid4()),
        "filename": filename,
        "sheet_name": hierarchy["flat_structure"][0]["sheet"] if hierarchy["flat_structure"] else "Unknown",
        "chunk_type": "document_overview",
        "document_context": doc_context,
        "sections": [
            {
                "code": s["code"],
                "title": s["title"],
                "subsections": [sub["title"] for sub in s.get("subsections", [])]
            }
            for s in hierarchy["sections"]
        ],
        "searchable_text": f"Document: {doc_context}\nSections: " +
                          ", ".join([f"{s['code']} {s['title']}" for s in hierarchy["sections"]]),
        "context_path": [doc_context]
    }
    chunks.append(overview_chunk)
   
    return chunks


def process_sheet_with_hierarchy(sheet_name: str, sheet_data: Dict[str, Any],
                                 filename: str) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    """
    Process a sheet to extract both traditional semantic chunks AND hierarchical structure.
    """
    rows = sheet_data.get("rows", [])
    max_col = sheet_data.get("max_col", 0)
   
    # 1. Build document hierarchy
    hierarchy = build_document_hierarchy(rows, sheet_name)
   
    # 2. Create context-aware chunks
    hierarchical_chunks = create_context_aware_chunks(hierarchy, filename)
   
    # 3. Traditional region detection
    regions = detect_table_regions(rows, max_col)
    tables = extract_structured_tables(rows, regions)
    traditional_chunks = create_semantic_chunks(tables, sheet_name, filename)
   
    # 4. Combine both
    all_chunks = hierarchical_chunks + traditional_chunks
   
    enriched_sheet = {
        "max_row": sheet_data.get("max_row", 0),
        "max_col": max_col,
        "rows": rows,
        "hierarchy": hierarchy,
        "regions": regions,
        "tables": tables,
        "chunks": all_chunks
    }
   
    return enriched_sheet, all_chunks


# ==================== KNOWLEDGE GRAPH CONSTRUCTION (NEO4J) ====================

async def create_neo4j_constraints():
    """Create constraints and indexes in Neo4j for better performance."""
    async with neo4j_driver.session() as session:
        constraints = [
            "CREATE CONSTRAINT document_id IF NOT EXISTS FOR (d:Document) REQUIRE d.id IS UNIQUE",
            "CREATE CONSTRAINT item_name IF NOT EXISTS FOR (i:Item) REQUIRE i.name IS UNIQUE",
            "CREATE CONSTRAINT sheet_id IF NOT EXISTS FOR (s:Sheet) REQUIRE s.id IS UNIQUE",
            "CREATE CONSTRAINT table_id IF NOT EXISTS FOR (t:Table) REQUIRE t.id IS UNIQUE",
            "CREATE INDEX item_type IF NOT EXISTS FOR (i:Item) ON (i.item_type)",
            "CREATE INDEX currency_index IF NOT EXISTS FOR (c:Cost) ON (c.currency)",
        ]
        
        for constraint in constraints:
            try:
                await session.run(constraint)
            except Exception as e:
                logger.debug(f"Constraint/Index already exists or failed: {e}")

async def build_knowledge_graph(document: Dict[str, Any], chunks: List[Dict[str, Any]]):
    async with neo4j_driver.session() as session:
        doc_id = document["_id"]
        filename = document["filename"]

        # 1. Document
        await session.run("""
            MERGE (d:Document {id: $doc_id})
            SET d.filename = $filename, d.uploaded_at = $uploaded_at
        """, doc_id=doc_id, filename=filename, uploaded_at=document["uploaded_at"].isoformat())

        # 2. Sheets & Tables
        for sheet_name, sheet_data in document["sheets"].items():
            sheet_id = f"{doc_id}_{sheet_name}"
            await session.run("""
                MERGE (s:Sheet {id: $sheet_id})
                SET s.name = $sheet_name
                MERGE (d:Document {id: $doc_id})
                MERGE (d)-[:CONTAINS_SHEET]->(s)
            """, sheet_id=sheet_id, sheet_name=sheet_name, doc_id=doc_id)

            for table in sheet_data.get("tables", []):
                table_id = f"{sheet_id}_table_{table['region_id']}"
                await session.run("""
                    MERGE (t:Table {id: $table_id})
                    SET t.start_row = $start, t.end_row = $end
                    MERGE (s:Sheet {id: $sheet_id})
                    MERGE (s)-[:CONTAINS_TABLE]->(t)
                """, table_id=table_id, start=table['start_row'], end=table['end_row'], sheet_id=sheet_id)

                await extract_entities_from_table(session, table_id, table, doc_id)

async def llm_infer_column_mappings(headers: List[str], sample_rows: List[List[Any]]) -> Dict[str, int]:
    """Use OpenAI LLM to intelligently infer column mappings."""
    import openai
    
    openai.api_key = os.getenv("OPENAI_API_KEY")
    
    sample_data = "\n".join([
        f"Row {i}: {dict(zip(headers, row))}"
        for i, row in enumerate(sample_rows[:3])
    ])
    
    prompt = f"""You are analyzing a construction BOQ (Bill of Quantities) Excel table.

Headers: {headers}

Sample Data:
{sample_data}

Identify which columns correspond to these semantic roles:
- item_description: Column containing item names/descriptions
- quantity: Column containing numeric quantities
- unit: Column containing measurement units
- rate: Column containing unit rates/prices
- amount: Column containing total amounts/costs
- code: Column containing item codes or S.No

Return ONLY a JSON object mapping role names to column indices (0-based). If a role doesn't exist, use -1.
Example: {{"item_description": 2, "quantity": 5, "unit": 6, "rate": 7, "amount": 8, "code": 0}}

JSON Response:"""

    try:
        response = openai.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are a BOQ analysis expert. Return only valid JSON."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.1,
            max_tokens=200
        )
        
        result_text = response.choices[0].message.content.strip()
        if result_text.startswith("```"):
            result_text = result_text.split("```")[1]
            if result_text.startswith("json"):
                result_text = result_text[4:]
        result_text = result_text.strip()
        
        mappings = json.loads(result_text)
        logger.info(f"LLM inferred column mappings: {mappings}")
        return mappings
    except Exception as e:
        logger.error(f"LLM column mapping failed: {e}, falling back to heuristics")
        return None


async def extract_entities_from_table(session, table_id: str, table: Dict[str, Any], doc_id: str):
    headers = table.get('headers', [])
    raw_rows = table.get('rows', [])
    enriched_rows = table.get('enriched_rows', [])

    if not raw_rows:
        return

    column_map = await llm_infer_column_mappings(headers, raw_rows[:5])
    if not column_map:
        logger.warning(f"LLM failed for table {table_id}, skipping graph")
        return

    for row_idx, raw_row in enumerate(raw_rows):
        row_id = f"{table_id}_row_{row_idx}"
        enriched_row = enriched_rows[row_idx] if row_idx < len(enriched_rows) else []

        item_text = None
        item_col = column_map.get('item_description', -1)
        if item_col >= 0 and item_col < len(raw_row):
            item_text = str(raw_row[item_col]).strip()
        if not item_text or item_text in ['nan', 'None', '']:
            continue

        await session.run("""
            MERGE (ii:ItemInstance {id: $row_id})
            SET ii.description = $desc,
                ii.document_id = $doc_id,
                ii.table_id = $table_id,
                ii.row_index = $row_idx
            WITH ii
            MATCH (t:Table {id: $table_id})
            MERGE (t)-[:HAS_ITEM]->(ii)
        """, row_id=row_id, desc=item_text, doc_id=doc_id, table_id=table_id, row_idx=row_idx)

        for role, col_idx in column_map.items():
            if col_idx < 0 or col_idx >= len(raw_row):
                continue
            value = raw_row[col_idx]
            if value is None or str(value).strip() in ['', 'nan']:
                continue

            if role in ['quantity', 'rate', 'amount']:
                try:
                    num_val = float(str(value).replace(',', ''))
                except:
                    continue
                await session.run(f"""
                    MATCH (ii:ItemInstance {{id: $row_id}})
                    MERGE (ii)-[:HAS_{role.upper()}]->(n:Number {{value: $val}})
                    SET n.unit = $unit, n.currency = $currency
                """, row_id=row_id, val=num_val,
                     unit=extract_unit(value), currency=extract_currency(value))
            elif role == 'unit':
                unit = normalize_unit_value(value)
                if unit:
                    await session.run("""
                        MATCH (ii:ItemInstance {id: $row_id})
                        MERGE (ii)-[:HAS_UNIT]->(u:Unit {name: $unit})
                    """, row_id=row_id, unit=unit)
            elif role == 'code':
                await session.run("""
                    MATCH (ii:ItemInstance {id: $row_id})
                    SET ii.code = $code
                """, row_id=row_id, code=str(value).strip())

def extract_unit(value) -> Optional[str]:
    if not isinstance(value, str):
        return None
    match = re.search(r'([a-zA-Z²³³]+)', str(value))
    return match.group(1) if match else None

def extract_currency(value) -> Optional[str]:
    if not isinstance(value, str):
        return None
    for pattern, curr in CURRENCY_PATTERNS:
        if re.search(pattern, value):
            return curr
    return None

def normalize_unit_value(value) -> Optional[str]:
    if not isinstance(value, str):
        return None
    val = value.lower().strip()
    return UNIT_MAPPINGS.get(val, val)

async def query_knowledge_graph_stats(doc_id: str) -> Dict[str, Any]:
    async with neo4j_driver.session() as session:
        result = await session.run("""
            MATCH (d:Document {id: $doc_id})
            OPTIONAL MATCH (d)-[:CONTAINS_SHEET]->(s)
            OPTIONAL MATCH (s)-[:CONTAINS_TABLE]->(t)
            OPTIONAL MATCH (t)-[:HAS_ITEM]->(ii:ItemInstance)
            OPTIONAL MATCH (ii)-[:HAS_QUANTITY]->(q:Number)
            OPTIONAL MATCH (ii)-[:HAS_AMOUNT]->(a:Number)
            OPTIONAL MATCH (ii)-[:HAS_RATE]->(r:Number)
            RETURN 
                count(DISTINCT s) as sheets,
                count(DISTINCT t) as tables,
                count(DISTINCT ii) as items,
                count(DISTINCT q) as quantities,
                count(DISTINCT a) as amounts,
                count(DISTINCT r) as rates
        """, doc_id=doc_id)
        rec = await result.single()
        return {
            "sheets": rec["sheets"],
            "tables": rec["tables"],
            "items": rec["items"],
            "quantities": rec["quantities"],
            "amounts": rec["amounts"],
            "rates": rec["rates"]
        }

# --- Helpers ---
def excel_serial_to_datetime(n: float) -> str | None:
    try:
        dt = EXCEL_EPOCH + timedelta(days=float(n))
        return dt.isoformat()
    except:
        return None

def col_letters_to_index(letters: str) -> int:
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch.upper()) - ord('A') + 1)
    return n

def cell_ref_to_coord(cell_ref: str) -> tuple:
    m = re.match(r"^([A-Za-z]+)(\d+)$", cell_ref)
    if not m:
        return 0, 0
    col_letters, row_num = m.groups()
    return int(row_num), col_letters_to_index(col_letters)

# --- XLSX Zip Parser ---
class XLSXZipParser:
    def __init__(self, data: bytes):
        self.zf = zipfile.ZipFile(io.BytesIO(data))
        self.shared_strings = []
        self.num_fmts = {}
        self.cell_xfs_numfmt = []
        self.sheet_map = []
        self._load_shared_strings()
        self._load_styles()
        self._load_sheet_map()

    def _load_shared_strings(self):
        try:
            with self.zf.open("xl/sharedStrings.xml") as f:
                root = ET.parse(f).getroot()
                for si in root.findall("s:si", NS):
                    parts = [t.text or "" for t in si.findall("s:t", NS)]
                    self.shared_strings.append("".join(parts))
        except KeyError:
            self.shared_strings = []

    def _load_styles(self):
        try:
            with self.zf.open("xl/styles.xml") as f:
                root = ET.parse(f).getroot()
                for nf in root.findall("s:numFmts/s:numFmt", NS):
                    numFmtId = int(nf.attrib.get("numFmtId", "-1"))
                    self.num_fmts[numFmtId] = nf.attrib.get("formatCode", "")
                self.cell_xfs_numfmt = [
                    int(xf.attrib.get("numFmtId", "-1"))
                    for xf in root.findall("s:cellXfs/s:xf", NS)
                ]
        except KeyError:
            pass

    def _load_sheet_map(self):
        try:
            with self.zf.open("xl/workbook.xml") as f:
                wb_root = ET.parse(f).getroot()
                sheets = [
                    (s.attrib.get("name", "Sheet"), s.attrib.get(f"{{{NS_REL}}}id"))
                    for s in wb_root.findall("s:sheets/s:sheet", NS)
                    if s.attrib.get(f"{{{NS_REL}}}id")
                ]
            with self.zf.open("xl/_rels/workbook.xml.rels") as f:
                rels_root = ET.parse(f).getroot()
                rel_map = {r.attrib["Id"]: r.attrib["Target"] for r in rels_root.findall("p:Relationship", NS)}
            self.sheet_map = [
                (name, "xl/" + rel_map.get(rid, "").lstrip("xl/"))
                for name, rid in sheets if rid in rel_map
            ]
        except Exception:
            self.sheet_map = []

    def _is_date_style(self, s_idx: int) -> bool:
        if s_idx < 0 or s_idx >= len(self.cell_xfs_numfmt):
            return False
        fmt_id = self.cell_xfs_numfmt[s_idx]
        if fmt_id in BUILTIN_DATE_FMT_IDS:
            return True
        return bool(re.search(r"[dyhmSs]", self.num_fmts.get(fmt_id, ""), re.I))

    def _apply_merged_cells(self, root: ET.Element, grid: dict):
        for merge in root.findall("s:mergeCells/s:mergeCell", NS):
            ref = merge.attrib.get("ref", "")
            m = re.match(r"^([A-Z]+\d+):([A-Z]+\d+)$", ref)
            if not m: continue
            tl, br = m.groups()
            r1, c1 = cell_ref_to_coord(tl)
            r2, c2 = cell_ref_to_coord(br)
            val = grid.get(r1, {}).get(c1)
            for r in range(r1, r2 + 1):
                row_map = grid.setdefault(r, {})
                for c in range(c1, c2 + 1):
                    if (r, c) != (r1, c1):
                        row_map[c] = val

    def parse_sheet(self, path: str) -> Dict[str, Any]:
        try:
            with self.zf.open(path) as f:
                root = ET.parse(f).getroot()
        except Exception:
            return {"max_row": 0, "max_col": 0, "rows": []}

        grid = {}
        max_row = max_col = 0
        for row in root.findall("s:sheetData/s:row", NS):
            r_idx = int(row.attrib.get("r", "0"))
            max_row = max(max_row, r_idx)
            for c in row.findall("s:c", NS):
                ref = c.attrib.get("r")
                if not ref: continue
                rr, cc = cell_ref_to_coord(ref)
                max_col = max(max_col, cc)
                t = c.attrib.get("t")
                s_attr = c.attrib.get("s")
                style_idx = int(s_attr) if s_attr and s_attr.isdigit() else -1
                v = c.find("s:v", NS)
                is_el = c.find("s:is", NS)
                val = None
                if t == "inlineStr" and is_el is not None:
                    tt = is_el.find("s:t", NS)
                    val = tt.text if tt is not None else None
                elif v is not None and v.text is not None:
                    text = v.text.strip()
                    if t == "s":
                        idx = int(text) if text.isdigit() else -1
                        val = self.shared_strings[idx] if 0 <= idx < len(self.shared_strings) else text
                    elif t == "b":
                        val = text == "1"
                    else:
                        if self._is_date_style(style_idx):
                            val = excel_serial_to_datetime(float(text)) or text
                        else:
                            try:
                                val = float(text) if "." in text or "e" in text.lower() else int(text)
                            except:
                                val = text
                grid.setdefault(rr, {})[cc] = val

        self._apply_merged_cells(root, grid)
        rows = [[grid.get(r, {}).get(c) for c in range(1, max_col + 1)] for r in range(1, max_row + 1)]
        return {"max_row": max_row, "max_col": max_col, "rows": rows}

    def parse_workbook(self) -> Dict[str, Any]:
        result = {"sheets": {}, "sheet_count": len(self.sheet_map)}
        for name, path in self.sheet_map:
            result["sheets"][name] = self.parse_sheet(path)
        return result


# === High-level parser ===
def parse_excel_bytes(data: bytes) -> Dict[str, Any]:
    if pd is not None:
        try:
            xl = pd.ExcelFile(io.BytesIO(data))
            result = {"sheets": {}, "sheet_count": len(xl.sheet_names)}
            for sheet in xl.sheet_names:
                df = pd.read_excel(io.BytesIO(data), sheet_name=sheet, dtype=object, engine='openpyxl')
                df = df.where(pd.notnull(df), None)
                matrix = df.values.tolist()
                result["sheets"][sheet] = {
                    "max_row": len(matrix),
                    "max_col": len(matrix[0]) if matrix else 0,
                    "rows": matrix,
                }
            return result
        except Exception as e:
            logger.debug(f"pandas failed: {e}")

    if load_workbook is not None:
        try:
            wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
            result = {"sheets": {}, "sheet_count": len(wb.sheetnames)}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for r in ws.iter_rows(values_only=True):
                    row = [v.isoformat() if isinstance(v, (datetime.datetime, datetime.date)) else v for v in r]
                    rows.append(row)
                max_row = ws.max_row or 0
                max_col = ws.max_column or 0
                result["sheets"][sheet_name] = {"max_row": max_row, "max_col": max_col, "rows": rows}
            return result
        except Exception as e:
            logger.debug(f"openpyxl failed: {e}")

    parser = XLSXZipParser(data)
    return parser.parse_workbook()


# === ENHANCED SINGLE FILE PROCESSOR ===
def process_single_file(file: UploadFile, user_id: str = None) -> Dict[str, Any]:
    """
    Enhanced version with hierarchical analysis.
    """
    content = file.file.read()
    if not content:
        raise ValueError("Empty file")
    parsed = parse_excel_bytes(content)
    doc_id = str(uuid.uuid4())
    now = datetime.datetime.utcnow()
    all_chunks = []
    enriched_sheets = {}
   
    for sheet_name, sheet_data in parsed.get("sheets", {}).items():
        enriched_sheet, chunks = process_sheet_with_hierarchy(sheet_name, sheet_data, file.filename)
        all_chunks.extend(chunks)
        enriched_sheets[sheet_name] = enriched_sheet

    document = {
        "_id": doc_id,
        "filename": file.filename,
        "user_id": user_id,
        "uploaded_at": now,
        "parsed_at": now,
        "sheet_count": parsed.get("sheet_count", 0),
        "sheets": enriched_sheets,
        "status": "parsed",
        "size_bytes": len(content),
        "processing_metadata": {
            "has_currency_data": any(c.get('metadata', {}).get('has_currency', False) for c in all_chunks),
            "has_quantity_data": any(c.get('metadata', {}).get('has_quantities', False) for c in all_chunks),
            "has_item_descriptions": any(c.get('metadata', {}).get('has_items', False) for c in all_chunks),
            "has_hierarchical_structure": any(c.get('chunk_type') == 'hierarchical_section' for c in all_chunks),
            "total_sections": sum(len(s.get('hierarchy', {}).get('sections', [])) for s in enriched_sheets.values()),
            "total_tables": sum(len(s['tables']) for s in enriched_sheets.values()),
            "total_regions": sum(len(s['regions']) for s in enriched_sheets.values())
        }
    }
    return {
        "document": document,
        "chunks": all_chunks,
        "document_id": doc_id,
        "filename": file.filename,
        "sheet_count": document["sheet_count"],
        "chunk_count": len(all_chunks),
        "status": "success",
        "uploaded_at": now.isoformat(),
        "size_bytes": len(content),
        "metadata": document["processing_metadata"]
    }


async def embed_and_store_chunks(chunks: List[Dict[str, Any]]):
    if not chunks:
        return

    openai.api_key = os.getenv("OPENAI_API_KEY")
    if not openai.api_key:
        raise ValueError("OPENAI_API_KEY not set")

    texts = []
    points = []

    for chunk in chunks:
        if chunk['chunk_type'] == 'hierarchical_section':
            text = chunk['searchable_text']
        else:
            headers = " | ".join(chunk["headers"])
            sample_rows = "\n".join([
                " | ".join(map(str, row)) for row in chunk["sample_data"]["raw"]
            ])
            text = f"Table: {chunk['sheet_name']}\nHeaders: {headers}\nData:\n{sample_rows}"
        texts.append(text)

    try:
        response = openai.embeddings.create(
            input=texts,
            model=EMBEDDING_MODEL
        )
        embeddings = [r.embedding for r in response.data]
    except Exception as e:
        logger.error(f"Embedding failed: {e}")
        return

    for i, (chunk, embedding) in enumerate(zip(chunks, embeddings)):
        point = PointStruct(
            id=chunk["chunk_id"],
            vector=embedding,
            payload={
                "document_id": chunk.get("document_id", ""),
                "filename": chunk["filename"],
                "sheet_name": chunk["sheet_name"],
                "chunk_type": chunk["chunk_type"],
                "region_id": chunk.get("region_id"),
                "row_range": chunk.get("row_range"),
                "headers": chunk.get("headers", []),
                "has_currency": chunk.get("metadata", {}).get("has_currency", False),
                "has_quantities": chunk.get("metadata", {}).get("has_quantities", False),
                "has_items": chunk.get("metadata", {}).get("has_items", False),
                "currency_columns": chunk.get("metadata", {}).get("currency_columns", []),
                "quantity_columns": chunk.get("metadata", {}).get("quantity_columns", []),
                "item_columns": chunk.get("metadata", {}).get("item_columns", []),
                "context_path": chunk.get("context_path", []),
                "section_code": chunk.get("section_code"),
                "section_title": chunk.get("section_title"),
            }
        )
        points.append(point)

    qdrant_client.upsert(
        collection_name=QDRANT_COLLECTION,
        points=points
    )
    logger.info(f"Stored {len(points)} chunk embeddings in Qdrant")


async def embed_and_store_document(document: Dict[str, Any]):
    openai.api_key = os.getenv("OPENAI_API_KEY")
    if not openai.api_key:
        raise ValueError("OPENAI_API_KEY not set")

    sheet_names = list(document["sheets"].keys())
    searchable_text = (
        f"File: {document['filename']}\n"
        f"Sheets: {', '.join(sheet_names)}\n"
        f"Total sheets: {document['sheet_count']}\n"
        f"Tables: {document['processing_metadata']['total_tables']}\n"
        f"Size: {document['size_bytes']} bytes"
    )

    resp = openai.embeddings.create(
        input=[searchable_text],
        model=EMBEDDING_MODEL,
    )
    embedding = resp.data[0].embedding

    payload = {
        **document,
        "_id": str(document["_id"]),
        "uploaded_at": document["uploaded_at"].isoformat(),
        "parsed_at": document["parsed_at"].isoformat(),
    }

    point = PointStruct(
        id=str(document["_id"]),
        vector=embedding,
        payload=payload,
    )

    qdrant_client.upsert(
        collection_name=QDRANT_COLLECTION_1,
        points=[point],
    )
    logger.info(f"Qdrant excel_documents point created (id={document['_id']})")


# === MAIN: PARALLEL INGESTION ===
async def ingest_excel_files(files: List[UploadFile], user_id: str = None) -> List[Dict[str, Any]]:
    if not files:
        raise ValueError("No files uploaded")

    try:
        await create_neo4j_constraints()
        logger.info("Neo4j constraints initialized")
    except Exception as e:
        logger.warning(f"Failed to create Neo4j constraints: {e}")

    results = []
    failed = []
    documents_to_insert = []
    chunks_to_insert = []

    future_to_file = {
        executor.submit(process_single_file, file, user_id): file
        for file in files
    }

    for future in as_completed(future_to_file):
        file = future_to_file[future]
        try:
            result = future.result()
            document = result["document"]
            chunks = result["chunks"]
            
            for chunk in chunks:
                chunk["document_id"] = document["_id"]
                chunks_to_insert.append(chunk)
            
            documents_to_insert.append(document)
            results.append({k: v for k, v in result.items() if k not in ["document", "chunks"]})
            logger.info(f"Parsed: {file.filename} → {result['document_id']} ({result['chunk_count']} chunks)")
        except Exception as e:
            error_msg = f"Failed to process {file.filename}: {str(e)}"
            logger.error(error_msg)
            failed.append({
                "filename": file.filename,
                "status": "failed",
                "error": str(e)
            })

        try:
            await embed_and_store_chunks(chunks)
            for result in results:
                if result["document_id"] == document["_id"]:
                    result["qdrant_status"] = "embedded"
                    break
        except Exception as e:
            logger.error(f"Qdrant upsert failed: {e}")
            for result in results:
                if result["document_id"] == document["_id"]:
                    result["qdrant_status"] = "failed"
                    result["qdrant_error"] = str(e)
                    break

    if documents_to_insert:
        try:
            collection = db["excel_documents"]
            await collection.insert_many(documents_to_insert, ordered=False)
            logger.info(f"Inserted {len(documents_to_insert)} documents to MongoDB")
            
            await init_qdrant_excel_collection()

            for doc in documents_to_insert:
                await embed_and_store_document(doc)

            if chunks_to_insert:
                chunks_collection = db["semantic_chunks"]
                await chunks_collection.insert_many(chunks_to_insert, ordered=False)
                logger.info(f"Inserted {len(chunks_to_insert)} semantic chunks to MongoDB")
        except Exception as e:
            logger.error(f"MongoDB batch insert failed: {str(e)}")
            for result in results:
                result["status"] = "failed"
                result["error"] = f"Database insert failed: {str(e)}"

    if documents_to_insert:
        for document in documents_to_insert:
            try:
                doc_id = document["_id"]
                doc_chunks = [c for c in chunks_to_insert if c.get("document_id") == doc_id]
                await build_knowledge_graph(document, doc_chunks)
                kg_stats = await query_knowledge_graph_stats(doc_id)
                for result in results:
                    if result.get("document_id") == doc_id:
                        result["knowledge_graph"] = kg_stats
                        break
                logger.info(f"Knowledge graph built for {document['filename']}: {kg_stats}")
            except Exception as e:
                logger.error(f"Failed to build knowledge graph for {document['filename']}: {e}")
                for result in results:
                    if result.get("document_id") == document["_id"]:
                        result["knowledge_graph_error"] = str(e)
                        break

    return results + failed


async def close_neo4j_driver():
    await neo4j_driver.close()
    logger.info("Neo4j driver closed")